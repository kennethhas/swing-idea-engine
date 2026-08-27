#!/usr/bin/env python3
"""
zone_scanner.py — deterministic supply/demand zone detector + Odds Enhancer scorer.

Implements the Seiden/OTA framework programmatically so zone detection is repeatable
instead of eyeballed. It is a FIRST PASS: the Big Picture, Arrival, and Curve enhancers
and any borderline base still deserve human/visual review. The script says so in its
output and flags low-confidence calls rather than hiding them.

Usage:
    python zone_scanner.py --ticker AAPL --timeframe daily
    python zone_scanner.py --csv data.csv --timeframe 1h
    python zone_scanner.py --ticker SPY --xlsx zones.xlsx

CSV format: header row with Date,Open,High,Low,Close[,Volume]  (Volume optional).

Data fetch (--ticker) uses Yahoo Finance's keyless chart endpoint (no API key) and supports
daily, weekly, and intraday timeframes. For unsupported feeds, pass your own --csv.
"""

import argparse
import csv
import re
import sys
import urllib.parse
from dataclasses import dataclass, field, asdict

try:
    import requests
except ImportError:
    requests = None


# ----------------------------- data loading -----------------------------

@dataclass
class Bar:
    date: str
    o: float
    h: float
    l: float
    c: float
    v: float = 0.0


def validate_bars(bars, source_note):
    """Sanity-check OHLC integrity before deriving trade levels from it.
    Drops structurally broken rows and returns (clean_bars, warnings)."""
    clean, warnings = [], []
    dropped = 0
    prev_c = None
    for b in bars:
        ok = (b.h >= b.l and b.h >= b.o and b.h >= b.c
              and b.l <= b.o and b.l <= b.c and b.o > 0 and b.c > 0)
        if not ok:
            dropped += 1
            continue
        # flag suspicious single-bar jumps (possible split / un-adjusted data)
        if prev_c and prev_c > 0:
            chg = abs(b.c - prev_c) / prev_c
            if chg > 0.5:
                warnings.append(f"large jump {prev_c:.2f}->{b.c:.2f} on {b.date} "
                                f"(possible split/adjustment; levels may be off)")
        prev_c = b.c
        clean.append(b)
    if dropped:
        warnings.append(f"dropped {dropped} structurally invalid OHLC row(s)")
    warnings.append(f"data source: {source_note} — single unofficial feed; cross-check levels before trading")
    return clean, warnings


def load_csv(path):
    bars = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        cols = {k.lower(): k for k in reader.fieldnames}
        need = ["date", "open", "high", "low", "close"]
        for n in need:
            if n not in cols:
                sys.exit(f"CSV missing required column '{n}'. Found: {reader.fieldnames}")
        for row in reader:
            try:
                bars.append(Bar(
                    date=str(row[cols["date"]]),
                    o=float(row[cols["open"]]), h=float(row[cols["high"]]),
                    l=float(row[cols["low"]]), c=float(row[cols["close"]]),
                    v=float(row[cols["volume"]]) if "volume" in cols and row[cols["volume"]] else 0.0,
                ))
            except (ValueError, TypeError):
                continue
    return bars


_TF_MAP = {
    "daily": ("1d", "2y"), "1d": ("1d", "2y"), "d": ("1d", "2y"),
    "weekly": ("1wk", "5y"), "1wk": ("1wk", "5y"), "w": ("1wk", "5y"),
    "60m": ("60m", "180d"), "1h": ("60m", "180d"), "hourly": ("60m", "180d"),
    "30m": ("30m", "60d"), "15m": ("15m", "60d"), "5m": ("5m", "60d"),
}


def fetch_yahoo(ticker, timeframe, interval=None, rng=None):
    """Fetch OHLC from Yahoo's keyless chart endpoint (no API key)."""
    if requests is None:
        sys.exit("requests not available; supply --csv instead.")
    # Allowlist the symbol charset (Yahoo symbols: letters, digits, . - ^ =).
    # Prevents path/param injection and use of the request as an egress channel.
    if not re.fullmatch(r"[A-Za-z0-9.\-^=]{1,15}", ticker or ""):
        sys.exit(f"Refusing suspicious ticker '{ticker}'. Use a plain symbol like AAPL, BRK-B, ^GSPC.")
    iv, rg = _TF_MAP.get(timeframe.lower(), ("1d", "2y"))
    iv = interval or iv
    rg = rng or rg
    sym = urllib.parse.quote(ticker, safe="")
    q = urllib.parse.urlencode({"range": rg, "interval": iv})
    import datetime
    headers = {"User-Agent": "Mozilla/5.0"}
    last_err = None
    for host in ("query1.finance.yahoo.com", "query2.finance.yahoo.com"):
        url = f"https://{host}/v8/finance/chart/{sym}?{q}"
        try:
            r = requests.get(url, headers=headers, timeout=30)
            data = r.json()
        except Exception as e:  # noqa
            last_err = e
            continue
        res = (data.get("chart") or {}).get("result")
        if not res:
            err = (data.get("chart") or {}).get("error")
            last_err = err or "empty result"
            continue
        res = res[0]
        ts = res.get("timestamp") or []
        q = ((res.get("indicators") or {}).get("quote") or [{}])[0]
        opens, highs = q.get("open", []), q.get("high", [])
        lows, closes, vols = q.get("low", []), q.get("close", []), q.get("volume", [])
        bars = []
        intraday = iv.endswith("m")
        for i, t in enumerate(ts):
            o, h, l, c = opens[i], highs[i], lows[i], closes[i]
            if None in (o, h, l, c):
                continue
            dt = datetime.datetime.fromtimestamp(t, datetime.timezone.utc)
            label = dt.strftime("%Y-%m-%d %H:%M" if intraday else "%Y-%m-%d")
            bars.append(Bar(label, float(o), float(h), float(l), float(c),
                            float(vols[i]) if i < len(vols) and vols[i] else 0.0))
        if bars:
            return bars
        last_err = "parsed 0 bars"
    sys.exit(f"Could not fetch '{ticker}' from Yahoo ({last_err}). Check the symbol or pass --csv.")


# ----------------------------- indicators -----------------------------

def atr(bars, period=14):
    """Wilder-ish ATR as a list aligned to bars (first `period` are seeded with simple mean)."""
    trs = []
    for i, b in enumerate(bars):
        if i == 0:
            trs.append(b.h - b.l)
        else:
            pc = bars[i - 1].c
            trs.append(max(b.h - b.l, abs(b.h - pc), abs(b.l - pc)))
    out = []
    run = None
    for i, tr in enumerate(trs):
        if i < period:
            run = sum(trs[: i + 1]) / (i + 1)
        else:
            run = (run * (period - 1) + tr) / period
        out.append(run)
    return out


def sma(bars, period):
    out = []
    for i in range(len(bars)):
        lo = max(0, i - period + 1)
        window = [b.c for b in bars[lo : i + 1]]
        out.append(sum(window) / len(window))
    return out


# ----------------------------- zone detection -----------------------------

@dataclass
class Zone:
    zone_type: str          # "demand" or "supply"
    pattern: str            # DBR / RBR / RBD / DBD
    base_start: int
    base_end: int           # index of last basing candle (the basing candle)
    legout_end: int
    proximal: float
    distal: float
    formed_date: str
    # scores
    strength: int = 0
    rr_score: int = 0
    big_picture: int = 0
    freshness: int = 0
    time_at_level: int = 0
    core_score: int = 0
    # plan
    target: float = None
    rr_ratio: float = None
    entry: float = None
    stop: float = None
    confidence: str = "Med"
    status: str = "live"          # live | active(in-zone now) | invalidated
    invalidated_date: str = None
    flags: list = field(default_factory=list)


def body_hi(b): return max(b.o, b.c)
def body_lo(b): return min(b.o, b.c)


def detect_zones(bars, atr_list, leg_mult=1.6, max_base=6, max_zones=12):
    """
    Walk forward. An explosive leg-out = a run of >=1 candle whose move is large vs ATR.
    The base = the small candle(s) immediately before it (range < ~ATR), up to max_base.
    Classify by leg-in vs leg-out direction.
    """
    n = len(bars)
    zones = []
    i = 2
    while i < n:
        a = atr_list[i] or 1e-9
        b = bars[i]
        body = abs(b.c - b.o)
        is_explosive = body >= leg_mult * a and body > (b.h - b.l) * 0.5
        if not is_explosive:
            i += 1
            continue

        legout_dir = "up" if b.c > b.o else "down"

        # extend leg-out through consecutive same-direction explosive/strong candles
        j = i
        while j + 1 < n:
            nb = bars[j + 1]
            nbody = abs(nb.c - nb.o)
            same = ("up" if nb.c > nb.o else "down") == legout_dir
            if same and nbody >= 0.6 * leg_mult * (atr_list[j + 1] or a):
                j += 1
            else:
                break
        legout_end = j

        # find base: small candles immediately before i, clustered near the basing candle
        base_end = i - 1
        base_start = base_end
        small_thresh = 1.0  # base candle range < ~1 ATR
        base_ref_c = bars[base_end].c  # the basing candle (last small candle before leg-out)
        k = base_end
        count = 0
        while k >= 1 and count < max_base:
            rng = bars[k].h - bars[k].l
            tight = rng <= small_thresh * (atr_list[k] or a)
            # must stay clustered with the basing candle, else it's a separate leg, not the base
            clustered = abs(bars[k].c - base_ref_c) <= 1.2 * a
            if tight and clustered:
                base_start = k
                k -= 1
                count += 1
            else:
                break
        if count == 0:
            # no quiet base candle; the explosive candle has no clean pause before it
            i = legout_end + 1
            continue

        # leg-in direction: compare price a few candles before the base to the base
        look = max(0, base_start - 3)
        pre = bars[look].c
        base_ref = (bars[base_start].c + bars[base_end].c) / 2
        legin_dir = "down" if base_ref < pre else "up"

        if legout_dir == "up":
            zone_type = "demand"
            pattern = "DBR" if legin_dir == "down" else "RBR"
            proximal = max(body_hi(bars[t]) for t in range(base_start, base_end + 1))
            distal = min(bars[t].l for t in range(base_start, base_end + 1))
        else:
            zone_type = "supply"
            pattern = "RBD" if legin_dir == "up" else "DBD"
            proximal = min(body_lo(bars[t]) for t in range(base_start, base_end + 1))
            distal = max(bars[t].h for t in range(base_start, base_end + 1))

        zones.append(Zone(
            zone_type=zone_type, pattern=pattern,
            base_start=base_start, base_end=base_end, legout_end=legout_end,
            proximal=round(proximal, 4), distal=round(distal, 4),
            formed_date=bars[base_end].date,
        ))
        i = legout_end + 1

    # keep the most recent zones (closest in time to "now")
    zones = zones[-max_zones:]
    return zones


# ----------------------------- scoring -----------------------------

def score_zones(bars, atr_list, sma50, sma200, zones):
    n = len(bars)
    last_close = bars[-1].c

    # ===== PASS 1: freshness + mitigation/status for EVERY zone first =====
    # R:R (pass 2) filters out invalidated opposing zones, so all statuses must
    # be finalized before any target is chosen.
    for z in zones:
        touches = 0
        in_zone_prev = False
        lo, hi = min(z.proximal, z.distal), max(z.proximal, z.distal)
        for t in range(z.legout_end + 1, n):
            bt = bars[t]
            in_zone = bt.l <= hi and bt.h >= lo
            if in_zone and not in_zone_prev:
                touches += 1
            in_zone_prev = in_zone
            broke = (bt.c < z.distal) if z.zone_type == "demand" else (bt.c > z.distal)
            if broke and z.status != "invalidated":
                z.status = "invalidated"
                z.invalidated_date = bt.date
        z.freshness = 2 if touches == 0 else 1 if touches == 1 else 0
        if touches >= 2:
            z.flags.append(f"tested {touches}x (orders likely consumed)")
        # price-side / current-state (catches gaps the close-loop missed)
        if z.status != "invalidated":
            if z.zone_type == "demand" and last_close < z.distal:
                z.status = "invalidated"; z.invalidated_date = bars[-1].date
            elif z.zone_type == "supply" and last_close > z.distal:
                z.status = "invalidated"; z.invalidated_date = bars[-1].date
            elif lo <= last_close <= hi:
                z.status = "active"
        if z.status == "invalidated":
            z.flags.append(f"INVALIDATED — price closed through the zone ({z.invalidated_date}); not a live setup")
        elif z.status == "active":
            z.flags.append("price is INSIDE the zone now — reacting/under test, not a clean pending entry")

    # ===== PASS 2: strength, time, big picture, R:R, plan, confidence =====
    for z in zones:
        a = atr_list[z.base_end] or 1e-9
        height = abs(z.proximal - z.distal) or 1e-9

        # --- Strength: leg-out travel vs ATR ---
        legout_move = abs(bars[z.legout_end].c - bars[z.base_end].c)
        s_ratio = legout_move / a
        z.strength = 2 if s_ratio >= 3 else 1 if s_ratio >= 1.5 else 0

        # --- Time at level: base candle count ---
        base_len = z.base_end - z.base_start + 1
        z.time_at_level = 1 if base_len <= 2 else 0
        if base_len > 4:
            z.flags.append(f"wide base ({base_len} candles)")

        # --- Big picture: trend alignment via CURRENT SMA50/200 (not trend-at-formation) ---
        up_trend = last_close > sma50[-1] and sma50[-1] >= sma200[-1]
        down_trend = last_close < sma50[-1] and sma50[-1] <= sma200[-1]
        if z.zone_type == "demand":
            z.big_picture = 2 if up_trend else 0 if down_trend else 1
        else:
            z.big_picture = 2 if down_trend else 0 if up_trend else 1
        z.flags.append("Big Picture = current-trend based; confirm trend-at-formation + Arrival/Curve visually")

        # --- Reward/Risk: nearest opposing LIVE zone as target, else recent swing extreme ---
        target, target_src = None, "opposing-zone"
        opposing = [o for o in zones if o.zone_type != z.zone_type and o.status != "invalidated"]
        if z.zone_type == "demand":
            above = [o for o in opposing if o.proximal > z.proximal]
            if above:
                target = min(o.proximal for o in above)
        else:
            below = [o for o in opposing if o.proximal < z.proximal]
            if below:
                target = max(o.proximal for o in below)
        if target is None:
            target_src = "recent-swing (estimate)"
            target = max(b.h for b in bars[-60:]) if z.zone_type == "demand" else min(b.l for b in bars[-60:])
        reward = abs(target - z.proximal)
        z.target = round(target, 4)
        z.rr_ratio = round(reward / height, 2)
        # Sanity-bound the ratio: an absurd R:R is an artifact (stale/far target), not an edge.
        absurd_rr = z.rr_ratio > 10
        if absurd_rr:
            z.flags.append(f"R:R {z.rr_ratio} implausibly high — target source '{target_src}' likely stale/far; treat R:R as UNVERIFIED")
            z.rr_score = 1  # don't reward an artifact with full marks
        else:
            z.rr_score = 2 if z.rr_ratio >= 3 else 1 if z.rr_ratio >= 2 else 0
        if target_src.startswith("recent-swing"):
            z.flags.append("target is a swing estimate (no opposing zone) — verify")
        if z.rr_score == 0:
            z.flags.append("poor R:R — opposing zone is close")

        # --- core total ---
        z.core_score = (z.strength + z.rr_score + z.big_picture
                        + z.freshness + z.time_at_level)

        # --- trade plan ---
        z.entry = z.proximal
        if z.zone_type == "demand":
            z.stop = round(z.distal - 0.1 * height, 4)
        else:
            z.stop = round(z.distal + 0.1 * height, 4)

        # --- confidence (status-aware; invalidated can never be a setup) ---
        if z.status == "invalidated":
            z.confidence = "Invalid"
        elif z.core_score >= 7 and z.rr_score == 2 and z.big_picture > 0 and z.status == "live" and not absurd_rr:
            z.confidence = "High"
        elif z.core_score >= 5:
            z.confidence = "Med"
        else:
            z.confidence = "Low"

    # rank: live/active setups first (by score), invalidated pushed to the bottom
    def rank_key(z):
        dead = 1 if z.status == "invalidated" else 0
        return (dead, -z.core_score, abs(z.proximal - last_close))
    zones.sort(key=rank_key)
    return zones


# ----------------------------- output -----------------------------

def print_table(bars, zones):
    last = bars[-1].c
    print(f"\nCurrent close: {last:.4f}   |   Bars analyzed: {len(bars)}   "
          f"|   Range: {bars[0].date} → {bars[-1].date}")
    print("Scores are a FIRST-PASS heuristic. Big Picture / Arrival / Curve need visual confirmation.\n")
    hdr = ["#", "Pattern", "Type", "Status", "Proximal", "Distal", "Str", "R:R", "Big", "Fresh",
           "Time", "Core/9", "Target", "RR", "Stop", "Conf"]
    print("  ".join(f"{h:>8}" for h in hdr))
    for idx, z in enumerate(zones, 1):
        row = [idx, z.pattern, z.zone_type, z.status, f"{z.proximal:.2f}", f"{z.distal:.2f}",
               z.strength, z.rr_score, z.big_picture, z.freshness, z.time_at_level,
               z.core_score, f"{z.target:.2f}", z.rr_ratio, f"{z.stop:.2f}", z.confidence]
        print("  ".join(f"{str(c):>8}" for c in row))
    print("\nFlags:")
    for idx, z in enumerate(zones, 1):
        if z.flags:
            print(f"  #{idx} {z.pattern} @ {z.proximal:.2f}: " + "; ".join(z.flags))


def _safe(v):
    """Neutralize spreadsheet formula injection: a string cell starting with
    = + - @ (or a control char) is prefixed with ' so Excel treats it as text."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@", "\t", "\r", "\n"):
        return "'" + v
    return v


def write_xlsx(path, bars, zones, ticker):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "S&D Zones"
    title = f"Supply/Demand Zones — {_safe(ticker or 'data')}  (current close {bars[-1].c:.4f}, as of {bars[-1].date})"
    ws["A1"] = _safe(title)
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "First-pass scan (Seiden/OTA). Invalidated zones are broken levels, not setups. Verify Big Picture/Arrival/Curve visually. Educational, not financial advice."
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    headers = ["Rank", "Pattern", "Type", "Status", "Proximal", "Distal", "Zone Height",
               "Strength/2", "R:R/2", "BigPicture/2", "Freshness/2", "Time/1",
               "Core/9", "Entry", "Stop", "Target", "R:R ratio", "Confidence",
               "Formed", "Flags / Verify"]
    hrow = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for idx, z in enumerate(zones, 1):
        r = hrow + idx
        vals = [idx, z.pattern, z.zone_type, z.status, z.proximal, z.distal,
                round(abs(z.proximal - z.distal), 4),
                z.strength, z.rr_score, z.big_picture, z.freshness, z.time_at_level,
                z.core_score, z.entry, z.stop, z.target, z.rr_ratio, z.confidence,
                z.formed_date, "; ".join(z.flags)]
        for c, v in enumerate(vals, 1):
            ws.cell(row=r, column=c, value=_safe(v))
        conf_cell = ws.cell(row=r, column=18)
        fill = {"High": "C6EFCE", "Med": "FFEB9C", "Low": "FFC7CE", "Invalid": "808080"}.get(z.confidence)
        if fill:
            conf_cell.fill = PatternFill("solid", fgColor=fill)
        scell = ws.cell(row=r, column=4)
        if z.status == "invalidated":
            scell.fill = PatternFill("solid", fgColor="808080")
            scell.font = Font(color="FFFFFF", bold=True)
        elif z.status == "active":
            scell.fill = PatternFill("solid", fgColor="FFD966")
        tcell = ws.cell(row=r, column=3)
        tcell.fill = PatternFill("solid", fgColor="DDEBF7" if z.zone_type == "demand" else "FCE4D6")

    widths = [6, 9, 8, 12, 10, 10, 11, 10, 7, 12, 11, 7, 8, 10, 10, 10, 9, 11, 12, 44]
    from openpyxl.utils import get_column_letter
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"
    wb.save(path)


# ----------------------------- main -----------------------------

def main():
    ap = argparse.ArgumentParser(description="Detect & score supply/demand zones (Seiden/OTA).")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--ticker", help="Symbol to fetch OHLC from Yahoo (e.g. AAPL, SPY).")
    src.add_argument("--csv", help="Path to OHLC CSV (Date,Open,High,Low,Close[,Volume]).")
    ap.add_argument("--timeframe", default="daily", help="daily/weekly/1h/30m/15m/5m. Sets Yahoo interval+range and is a label.")
    ap.add_argument("--interval", help="Override Yahoo interval (e.g. 1d,1wk,60m).")
    ap.add_argument("--range", dest="rng", help="Override Yahoo range (e.g. 2y,6mo,180d).")
    ap.add_argument("--leg-mult", type=float, default=1.6, help="Leg-out strength multiple vs ATR (default 1.6).")
    ap.add_argument("--max-base", type=int, default=6, help="Max candles allowed in a base (default 6).")
    ap.add_argument("--xlsx", help="Optional path to write an Excel workbook.")
    args = ap.parse_args()

    raw = (fetch_yahoo(args.ticker, args.timeframe, args.interval, args.rng)
           if args.ticker else load_csv(args.csv))
    src_note = f"Yahoo Finance ({args.ticker})" if args.ticker else f"CSV ({args.csv})"
    bars, warnings = validate_bars(raw, src_note)
    if len(bars) < 30:
        sys.exit(f"Need >=30 valid bars; got {len(bars)}.")
    if warnings:
        print("Data checks:")
        for w in warnings:
            print(f"  ! {w}")

    atr_list = atr(bars)
    sma50 = sma(bars, 50)
    sma200 = sma(bars, 200)
    zones = detect_zones(bars, atr_list, leg_mult=args.leg_mult, max_base=args.max_base)
    if not zones:
        print("No clean base+leg-out zones detected. Price may be mid-range / in equilibrium. "
              "Loosen --leg-mult or inspect the chart visually.")
        return
    zones = score_zones(bars, atr_list, sma50, sma200, zones)
    print_table(bars, zones)
    if args.xlsx:
        write_xlsx(args.xlsx, bars, zones, args.ticker)
        print(f"\nWrote workbook: {args.xlsx}")


if __name__ == "__main__":
    main()
